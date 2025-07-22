import json
import boto3
from datetime import datetime
import os
from typing import Dict, List, Any
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 로깅 설정
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# AWS 클라이언트 초기화
s3_client = boto3.client('s3')

# 환경 변수 - S3만 사용하므로 DynamoDB 관련 제거
BUCKET_NAME = os.environ.get('S3_BUCKET_NAME', 'competency-surveys')

def lambda_handler(event, context):
    """
    AI 역량평가 결과를 Excel 파일로 export하는 Lambda 함수
    """
    try:
        # 요청 파라미터 파싱
        logger.info(f"Received event: {json.dumps(event)}")
        
        # HTTP 요청의 body에서 파라미터 추출
        if 'body' in event:
            if isinstance(event['body'], str):
                body = json.loads(event['body'])
            else:
                body = event['body']
        else:
            body = event
        
        workspace_name = body.get('workspace_name')
        survey_name = body.get('survey_name')
        
        logger.info(f"Parsed parameters: workspace_name={workspace_name}, survey_name={survey_name}")
        
        if not workspace_name or not survey_name:
            return {
                'statusCode': 400,
                'body': json.dumps({
                    'error': 'workspace_name and survey_name are required',
                    'message': '워크스페이스명과 설문명이 필요합니다.'
                })
            }
        
        logger.info(f"Starting export for workspace: {workspace_name}, survey: {survey_name}")
        
        # 1. S3에서 설문 응답 JSON 데이터 조회
        survey_responses = get_survey_responses(workspace_name, survey_name)
        if not survey_responses:
            return {
                'statusCode': 404,
                'body': json.dumps({
                    'error': 'No responses found',
                    'message': '응답 데이터가 없습니다.'
                })
            }
        
        # 2. Excel 파일 생성
        excel_file_path = create_excel_file(workspace_name, survey_name, survey_responses)
        
        # 3. S3에 업로드
        s3_key = f"exports/{workspace_name}/{survey_name}/{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        upload_to_s3(excel_file_path, s3_key)
        
        # 4. 다운로드 URL 생성 (7일간 유효)
        download_url = generate_presigned_url(s3_key, expiration=604800)  # 7 days
        
        logger.info(f"Export completed successfully for workspace: {workspace_name}, survey: {survey_name}")
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'success': True,
                'message': 'Export 완료',
                'download_url': download_url,
                's3_key': s3_key,
                'workspace_name': workspace_name,
                'survey_name': survey_name,
                'total_responses': len(survey_responses)
            })
        }
        
    except Exception as e:
        logger.error(f"Export failed: {str(e)}")
        return {
            'statusCode': 500,
            'body': json.dumps({
                'error': 'Internal server error',
                'message': f'Export 처리 중 오류가 발생했습니다: {str(e)}'
            })
        }



def get_survey_responses(workspace_name: str, survey_name: str) -> List[Dict[str, Any]]:
    """S3에서 설문 응답 JSON 파일들 조회"""
    try:
        # S3 경로 구성: reports/{workspace_name}/{survey_name}/AI/
        s3_prefix = f"reports/{workspace_name}/{survey_name}/AI/"
        
        logger.info(f"S3에서 JSON 파일 조회 시작: s3://{BUCKET_NAME}/{s3_prefix}")
        
        # S3에서 해당 경로의 모든 JSON 파일 목록 조회
        response = s3_client.list_objects_v2(
            Bucket=BUCKET_NAME,
            Prefix=s3_prefix
        )
        
        if 'Contents' not in response:
            logger.warning(f"해당 경로에 파일이 없습니다: {s3_prefix}")
            return []
        
        json_responses = []
        
        # 각 JSON 파일을 다운로드하고 파싱
        for obj in response['Contents']:
            if obj['Key'].endswith('.json'):
                try:
                    # S3에서 JSON 파일 다운로드
                    file_response = s3_client.get_object(
                        Bucket=BUCKET_NAME,
                        Key=obj['Key']
                    )
                    
                    # JSON 파싱
                    json_content = json.loads(file_response['Body'].read().decode('utf-8'))
                    json_responses.append(json_content)
                    
                    logger.info(f"JSON 파일 로딩 완료: {obj['Key']}")
                    
                except Exception as file_error:
                    logger.error(f"JSON 파일 처리 실패 {obj['Key']}: {str(file_error)}")
                    continue
        
        logger.info(f"총 {len(json_responses)}개의 응답 데이터 로딩 완료")
        return json_responses
        
    except Exception as e:
        logger.error(f"S3에서 설문 응답 조회 실패: {str(e)}")
        return []

def create_excel_file(workspace_name: str, survey_name: str, responses: List[Dict[str, Any]]) -> str:
    """Excel 파일 생성 (openpyxl 사용)"""
    try:
        # 임시 파일 경로
        safe_workspace = workspace_name.replace('/', '_').replace(' ', '_')
        safe_survey = survey_name.replace('/', '_').replace(' ', '_')
        file_path = f"/tmp/export_{safe_workspace}_{safe_survey}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Excel 워크북 생성
        wb = Workbook()
        
        # 기본 시트 제거
        wb.remove(wb.active)
        
        # 1. 개요 시트
        create_summary_sheet_openpyxl(wb, workspace_name, survey_name, responses)
        
        # 2. 카테고리별 점수 시트 (메인 시트)
        create_category_scores_sheet_openpyxl(wb, responses)
        
        # 파일 저장
        wb.save(file_path)
        
        logger.info(f"Excel file created: {file_path}")
        return file_path
        
    except Exception as e:
        logger.error(f"Failed to create Excel file: {str(e)}")
        raise

def create_summary_sheet_openpyxl(wb: Workbook, workspace_name: str, survey_name: str, responses: List[Dict[str, Any]]):
    """개요 시트 생성 (openpyxl 사용)"""
    ws = wb.create_sheet(title='개요')
    
    # 평균 점수 계산
    total_scores = [r.get('aiAnalysis', {}).get('overallScore', 0) for r in responses if r.get('aiAnalysis')]
    avg_score = sum(total_scores) / len(total_scores) if total_scores else 0
    
    # 완료된 응답 수 (aiAnalysisStatus가 completed인 것들)
    completed_responses = [r for r in responses if r.get('aiAnalysisStatus') == 'completed']
    
    # 헤더 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 헤더 작성
    ws['A1'] = '항목'
    ws['B1'] = '값'
    ws['A1'].font = header_font
    ws['B1'].font = header_font
    ws['A1'].fill = header_fill
    ws['B1'].fill = header_fill
    
    # 데이터 작성
    data = [
        ['워크스페이스명', workspace_name],
        ['설문명', survey_name],
        ['총 응답자 수', len(responses)],
        ['완료된 응답 수', len(completed_responses)],
        ['완료율', f"{len(completed_responses) / len(responses) * 100:.1f}%" if responses else "0%"],
        ['평균 점수', f"{avg_score:.2f}"],
        ['Export 생성일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row_idx, (item, value) in enumerate(data, start=2):
        ws[f'A{row_idx}'] = item
        ws[f'B{row_idx}'] = value
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40

def create_respondents_sheet_openpyxl(wb: Workbook, responses: List[Dict[str, Any]]):
    """응답자 정보 시트 생성 (openpyxl 사용)"""
    ws = wb.create_sheet(title='응답자 정보')
    
    # 헤더 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 헤더 작성
    headers = ['이름', '소속', '전공', '나이', '이메일', '학력', '응답 제출일시', 'AI 분석 상태', 'AI 분석 완료일시', '전체 점수']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # 데이터 작성
    for row_idx, response in enumerate(responses, start=2):
        student_info = response.get('studentInfo', {})
        ws.cell(row=row_idx, column=1, value=student_info.get('name', ''))
        ws.cell(row=row_idx, column=2, value=student_info.get('organization', ''))
        ws.cell(row=row_idx, column=3, value=student_info.get('major', ''))
        ws.cell(row=row_idx, column=4, value=student_info.get('age', ''))
        ws.cell(row=row_idx, column=5, value=student_info.get('email', ''))
        ws.cell(row=row_idx, column=6, value=student_info.get('education', ''))
        ws.cell(row=row_idx, column=7, value=response.get('submittedAt', ''))
        ws.cell(row=row_idx, column=8, value=response.get('aiAnalysisStatus', ''))
        ws.cell(row=row_idx, column=9, value=response.get('aiAnalysis', {}).get('analyzedAt', ''))
        ws.cell(row=row_idx, column=10, value=response.get('aiAnalysis', {}).get('overallScore', ''))
    
    # 열 너비 자동 조정
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

def create_responses_sheet_openpyxl(wb: Workbook, responses: List[Dict[str, Any]]):
    """설문 응답 시트 생성 (openpyxl 사용)"""
    ws = wb.create_sheet(title='설문 응답')
    
    if not responses:
        return
    
    # 헤더 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 모든 질문 수집 (첫 번째 응답에서)
    first_response = responses[0]
    student_headers = ['이름', '소속', '전공', '나이', '이메일', '학력']
    questions = list(first_response.get('answers', {}).keys())
    all_headers = student_headers + questions
    
    # 헤더 작성
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # 데이터 작성
    for row_idx, response in enumerate(responses, start=2):
        student_info = response.get('studentInfo', {})
        answers = response.get('answers', {})
        
        # 학생 정보
        ws.cell(row=row_idx, column=1, value=student_info.get('name', ''))
        ws.cell(row=row_idx, column=2, value=student_info.get('organization', ''))
        ws.cell(row=row_idx, column=3, value=student_info.get('major', ''))
        ws.cell(row=row_idx, column=4, value=student_info.get('age', ''))
        ws.cell(row=row_idx, column=5, value=student_info.get('email', ''))
        ws.cell(row=row_idx, column=6, value=student_info.get('education', ''))
        
        # 설문 응답
        for col_idx, question in enumerate(questions, start=7):
            ws.cell(row=row_idx, column=col_idx, value=answers.get(question, ''))
    
    # 열 너비 조정
    for col_idx in range(1, len(all_headers) + 1):
        if col_idx <= 6:  # 학생 정보 열
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        else:  # 질문 열
            ws.column_dimensions[get_column_letter(col_idx)].width = 30

def create_category_scores_sheet_openpyxl(wb: Workbook, responses: List[Dict[str, Any]]):
    """카테고리별 점수 시트 생성 (openpyxl 사용) - 메인 시트"""
    ws = wb.create_sheet(title='카테고리별 점수')
    
    if not responses:
        return
    
    # 헤더 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font.color = "FFFFFF"
    
    # 기본 헤더
    base_headers = ['이름', '소속', '전공', '전체 점수']
    
    # 첫 번째 응답에서 카테고리 정보 추출
    categories = []
    if responses:
        first_response = responses[0]
        ai_analysis = first_response.get('aiAnalysis', {})
        category_scores = ai_analysis.get('categoryScores', [])
        categories = [cat.get('category', '') for cat in category_scores]
    
    # 카테고리별 헤더 생성 (점수, 백분율, 레벨)
    category_headers = []
    for category in categories:
        category_headers.extend([
            f'{category}_점수',
            f'{category}_백분율',
            f'{category}_레벨'
        ])
    
    all_headers = base_headers + category_headers
    
    # 헤더 작성
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 데이터 작성
    for row_idx, response in enumerate(responses, start=2):
        student_info = response.get('studentInfo', {})
        ai_analysis = response.get('aiAnalysis', {})
        category_scores = ai_analysis.get('categoryScores', [])
        
        # 기본 정보
        ws.cell(row=row_idx, column=1, value=student_info.get('name', ''))
        ws.cell(row=row_idx, column=2, value=student_info.get('organization', ''))
        ws.cell(row=row_idx, column=3, value=student_info.get('major', ''))
        ws.cell(row=row_idx, column=4, value=ai_analysis.get('overallScore', ''))
        
        # 카테고리별 점수
        col_idx = 5
        for category in categories:
            # 해당 카테고리 찾기
            category_data = next((cat for cat in category_scores if cat.get('category') == category), {})
            
            # 점수, 백분율, 레벨 입력
            ws.cell(row=row_idx, column=col_idx, value=category_data.get('score', ''))
            ws.cell(row=row_idx, column=col_idx + 1, value=f"{category_data.get('percentage', '')}%" if category_data.get('percentage') else '')
            ws.cell(row=row_idx, column=col_idx + 2, value=category_data.get('level', ''))
            
            col_idx += 3
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 12  # 이름
    ws.column_dimensions['B'].width = 25  # 소속
    ws.column_dimensions['C'].width = 15  # 전공
    ws.column_dimensions['D'].width = 12  # 전체 점수
    
    # 카테고리별 열 너비 조정
    col_idx = 5
    for category in categories:
        ws.column_dimensions[get_column_letter(col_idx)].width = 12      # 점수
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 12  # 백분율
        ws.column_dimensions[get_column_letter(col_idx + 2)].width = 10  # 레벨
        col_idx += 3
    
    # 행 높이 조정
    ws.row_dimensions[1].height = 25

def upload_to_s3(file_path: str, s3_key: str):
    """S3에 파일 업로드"""
    try:
        s3_client.upload_file(file_path, BUCKET_NAME, s3_key)
        logger.info(f"File uploaded to S3: s3://{BUCKET_NAME}/{s3_key}")
    except Exception as e:
        logger.error(f"Failed to upload to S3: {str(e)}")
        raise

def generate_presigned_url(s3_key: str, expiration: int = 3600) -> str:
    """S3 presigned URL 생성"""
    try:
        url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': BUCKET_NAME, 'Key': s3_key},
            ExpiresIn=expiration
        )
        return url
    except Exception as e:
        logger.error(f"Failed to generate presigned URL: {str(e)}")
        raise